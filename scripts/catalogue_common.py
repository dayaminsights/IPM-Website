"""Shared data-layer + drawing helpers for the IPM catalogue PDF builders.
Used by build-catalogue-pdf.py, build-catalogue-hindware.py, build-catalogue-jaquar.py.
"""
import os, re, io
from pathlib import Path
from reportlab.lib.colors import HexColor
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader
from PIL import Image
import openpyxl

FONT_DIR = Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts"
FONTS = {"SegoeUI": "segoeui.ttf", "SegoeUI-Semibold": "seguisb.ttf", "SegoeUI-Bold": "segoeuib.ttf"}

def register_fonts():
    for name, fname in FONTS.items():
        p = FONT_DIR / fname
        if not p.exists():
            raise FileNotFoundError(f"Required font missing: {p}")
        pdfmetrics.registerFont(TTFont(name, str(p)))

SHEET = "Item Master + Images"

def _clean(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None

def load_products(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.worksheets[0]
    rows = ws.iter_rows(min_row=2, values_only=True)
    # column indexes by header order (fixed schema)
    NAME, CODE, COLL, CAT, MRP, SITE = 1, 2, 3, 4, 5, 10
    out = []
    for r in rows:
        if r is None or all(c is None or str(c).strip()=="" for c in r):
            continue
        code = _clean(r[CODE]); name = _clean(r[NAME])
        if not code or not name:
            continue
        mrp_raw = r[MRP]
        try:
            mrp = int(float(mrp_raw)) if mrp_raw not in (None, "") else None
        except (ValueError, TypeError):
            mrp = None
        if mrp == 0:            # 0 is not a real price -> treat as missing
            mrp = None
        out.append({
            "item_code": code,
            "name": name,
            "collection": _clean(r[COLL]) or "Uncategorised",
            "category": _clean(r[CAT]) or "",
            "mrp": mrp,
            "image_filename": _clean(r[SITE]),
            "image_path": None,
        })
    return out

def _index_images(images_root):
    # filename -> list of every full path carrying that name (dupes across folders)
    idx = {}
    for dirpath, _, files in os.walk(images_root):
        for f in files:
            if f.lower().endswith((".png", ".jpg", ".jpeg", ".webp")):
                idx.setdefault(f, []).append(os.path.join(dirpath, f))
    return idx

def _collection_slug(name):
    """'Zenith Gunmetal Black Collection' -> 'zenith-gunmetal-black'."""
    s = (name or "").lower().replace("collection", "")
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")

def _choose_path(paths, collection):
    """Prefer the copy whose folder name matches the product's collection slug;
    fall back to the first path when nothing matches."""
    if len(paths) == 1:
        return paths[0]
    slug = _collection_slug(collection)
    for p in paths:
        if os.path.basename(os.path.dirname(p)) == slug:
            return p
    return paths[0]

def resolve_images(products, images_root):
    idx = _index_images(images_root)
    included, no_image, broken = [], [], []
    used_paths = set()
    for p in products:
        fn = p["image_filename"]
        if not fn:
            no_image.append(p); continue
        paths = idx.get(fn)
        if not paths:
            broken.append(p); continue
        chosen = _choose_path(paths, p["collection"])
        p["image_path"] = chosen
        used_paths.add(chosen)
        included.append(p)
    # Orphans = physical files never chosen. One basename entry per unreferenced
    # physical file so len(orphans) reflects the true physical count; the report
    # dedups basenames only for display.
    all_paths = [pp for paths in idx.values() for pp in paths]
    orphans = sorted(os.path.basename(pp) for pp in all_paths if pp not in used_paths)
    return included, no_image, broken, orphans

SIGNATURE_PREFIXES = ["Zenith", "Opell Prima", "Para"]

def order_collections(products):
    # preserve first-appearance order of collections
    seen = []
    buckets = {}
    for p in products:
        c = p["collection"]
        if c not in buckets:
            buckets[c] = []; seen.append(c)
        buckets[c].append(p)

    def rank(name):
        for i, pre in enumerate(SIGNATURE_PREFIXES):
            if name.startswith(pre):
                return (i, seen.index(name))
        return (len(SIGNATURE_PREFIXES), seen.index(name))

    ordered_names = sorted(seen, key=rank)
    return [(name, buckets[name]) for name in ordered_names]

def paginate(groups, per_page=12):
    pages = []
    for name, items in groups:
        for i in range(0, len(items), per_page):
            pages.append((name, items[i:i+per_page]))
    return pages

def _fmt_price(mrp):
    # 0 is not a real price -> treat like missing (IPM-style formatting: "₹ : n/-")
    return f"₹ : {mrp}/-" if mrp else "Price on request"

def write_report(report_path, total, included, no_image, broken, price_on_request, orphans):
    Path(report_path).parent.mkdir(parents=True, exist_ok=True)
    L = []
    L.append("# Catalogue Build Report\n")
    L.append(f"- Total rows: {total}")
    L.append(f"- Included: {included}")
    L.append(f"- Omitted (no image): {len(no_image)}")
    L.append(f"- Omitted (broken image ref): {len(broken)}")
    L.append(f"- Price on request: {len(price_on_request)}")
    L.append(f"- Orphan photos (on disk, no included row): {len(orphans)}\n")

    def section(title, rows, fmt):
        L.append(f"\n## {title} ({len(rows)})\n")
        if not rows: L.append("_none_"); return
        for r in rows: L.append(fmt(r))

    section("Omitted — no image", no_image, lambda r: f"- {r['item_code']} — {r['name']} ({r['collection']})")
    section("Omitted — broken image reference", broken, lambda r: f"- {r['item_code']} — {r['name']} → {r.get('image_filename')}")
    section("Price on request", price_on_request, lambda r: f"- {r['item_code']} — {r['name']} ({r['collection']})")
    # dedup basenames for display; the summary count above stays physical (len(orphans))
    orphan_display = list(dict.fromkeys(orphans))
    section("Orphan photos", [{"f":o} for o in orphan_display], lambda r: f"- {r['f']}")

    Path(report_path).write_text("\n".join(L), encoding="utf-8")
    print(f"Report: {report_path}  |  included={included} no_image={len(no_image)} broken={len(broken)} price_req={len(price_on_request)} orphans={len(orphans)}")

def _draw_logo(c, x, y, scale=1.0, color="#111111"):
    """Minimal IPM wordmark: 'IPM' + 4 right bars + 'BATH FITTINGS'."""
    col = HexColor(color)
    c.setFillColor(col)
    c.setFont("SegoeUI-Bold", 26*scale)
    c.drawString(x, y, "IPM")
    bx = x + 52*scale
    c.setLineWidth(2.2*scale); c.setStrokeColor(col)
    for i, w in enumerate([26, 22, 18, 14]):
        yy = y + 18*scale - i*5.5*scale
        c.line(bx + (26-w)*scale, yy, bx + 26*scale, yy)
    c.setFont("SegoeUI", 6.5*scale)
    c.drawString(x, y - 9*scale, "B A T H   F I T T I N G S")

def _desaturate(path):
    im = Image.open(path).convert("RGB")
    g = im.convert("L").convert("RGB")
    buf = io.BytesIO(); g.save(buf, format="JPEG", quality=88); buf.seek(0)
    return ImageReader(buf)

# Cache downscaled product images by absolute source path so the same file is
# only decoded/resized once even if it appears on multiple pages (or in both
# a Hindware-style and Jaquar-style build run back to back).
_PRODUCT_IMG_CACHE = {}
PRODUCT_IMG_MAX_PX = 320

def _load_product_image(path):
    """Return (ImageReader, (w, h)) for a downscaled, transparency-preserving copy
    of the product photo. Cached by absolute path."""
    key = os.path.abspath(path)
    cached = _PRODUCT_IMG_CACHE.get(key)
    if cached is not None:
        return cached
    im = Image.open(path)
    im = im.convert("RGBA")  # keep transparency
    im.thumbnail((PRODUCT_IMG_MAX_PX, PRODUCT_IMG_MAX_PX), Image.LANCZOS)
    buf = io.BytesIO(); im.save(buf, format="PNG", optimize=True); buf.seek(0)
    reader = ImageReader(buf)
    result = (reader, im.size)
    _PRODUCT_IMG_CACHE[key] = result
    return result

def _wrap(c, text, font, size, max_w):
    words = text.split(); lines=[]; cur=""
    for wd in words:
        t = (cur+" "+wd).strip()
        if c.stringWidth(t, font, size) <= max_w: cur=t
        else: lines.append(cur); cur=wd
    if cur: lines.append(cur)
    return lines or [""]

# --- Collection hero-image resolution (for the Hindware/Jaquar-style collection
# opening pages, which use a real lifestyle/product photo per collection instead
# of a plain colored header band). ---

def _normalize_collection_filename(fname):
    """'Cube_Prima.jpg' -> 'cube-prima'; 'line-opell-prima.jpg' -> 'opell-prima';
    'NEO_COLLECTION.png' -> 'neo'."""
    stem = os.path.splitext(fname)[0]
    s = stem.lower().replace("_", "-").replace(" ", "-")
    s = re.sub(r"[^a-z0-9-]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    if s.startswith("line-"):
        s = s[len("line-"):]
    if s.endswith("-collection"):
        s = s[: -len("-collection")]
    return s

def _index_collection_heroes(images_collections_dir):
    idx = {}
    for f in sorted(os.listdir(images_collections_dir)):
        full = os.path.join(images_collections_dir, f)
        if not os.path.isfile(full):
            continue
        low = f.lower()
        if low.startswith("cat-") or low == "hero.jpg":
            continue
        if not low.endswith((".png", ".jpg", ".jpeg", ".webp")):
            continue
        key = _normalize_collection_filename(f)
        idx.setdefault(key, full)  # first match wins
    return idx

def _category_fallback_path(images_collections_dir, category):
    c = (category or "").lower()
    if "shower" in c:
        candidates = ["cat-shower.jpg", "cat-showers.jpg"]
    elif "mixer" in c or "kitchen" in c:
        candidates = ["cat-kitchen-mixers.jpg", "cat-mixers.jpg"]
    elif "accessor" in c:
        candidates = ["cat-accessories.jpg"]
    else:
        candidates = ["cat-faucets.jpg"]
    for name in candidates:
        p = os.path.join(images_collections_dir, name)
        if os.path.exists(p):
            return p
    return None

def resolve_collection_hero(collection_name, images_collections_dir, category, included_products_for_collection):
    """3-step fallback: (1) a known hero file matching the collection name
    (progressively shortened, e.g. 'zenith-gunmetal-black' -> 'zenith-gunmetal'
    -> 'zenith'), (2) the first included product's own photo, (3) the category
    fallback image. Returns an absolute path or None."""
    idx = _index_collection_heroes(images_collections_dir)
    slug = _collection_slug(collection_name)
    segments = [s for s in slug.split("-") if s]
    for n in range(len(segments), 0, -1):
        candidate = "-".join(segments[:n])
        if candidate in idx:
            return idx[candidate]
    for p in included_products_for_collection:
        if p.get("image_path") and os.path.exists(p["image_path"]):
            return p["image_path"]
    return _category_fallback_path(images_collections_dir, category)
