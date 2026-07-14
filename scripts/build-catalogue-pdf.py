"""Repeatable IPM product catalogue PDF builder. See docs/superpowers/specs/2026-07-14-product-catalogue-pdf-design.md"""
import os, sys
from pathlib import Path
from reportlab.lib.colors import HexColor
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

PAGE_W, PAGE_H = 612, 810
TEAL = HexColor("#2E9AA6")
GRAY_TAB = HexColor("#DADADA")
GRAY_TAB_TEXT = HexColor("#6E6E6E")
COLS, ROWS = 3, 4
HEADER_H, FOOTER_H = 96, 30
MARGIN_X = 40

FONT_DIR = Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts"
FONTS = {"SegoeUI": "segoeui.ttf", "SegoeUI-Semibold": "seguisb.ttf", "SegoeUI-Bold": "segoeuib.ttf"}

def register_fonts():
    for name, fname in FONTS.items():
        p = FONT_DIR / fname
        if not p.exists():
            raise FileNotFoundError(f"Required font missing: {p}")
        pdfmetrics.registerFont(TTFont(name, str(p)))

import openpyxl

SHEET = "Item Master + Images"

def _clean(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None

def load_products(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.worksheets[0]
    rows = ws.iter_rows(min_row=2, values_only=True)
    # column indexes by header order (fixed schema)
    NAME, CODE, COLL, CAT, MRP, SITE = 1, 2, 3, 4, 5, 10
    out = []
    for r in rows:
        if r is None or all(c is None or str(c).strip()=="" for c in r):
            continue
        code = _clean(r[CODE]); name = _clean(r[NAME])
        if not code or not name:
            continue
        mrp_raw = r[MRP]
        try:
            mrp = int(float(mrp_raw)) if mrp_raw not in (None, "") else None
        except (ValueError, TypeError):
            mrp = None
        out.append({
            "item_code": code,
            "name": name,
            "collection": _clean(r[COLL]) or "Uncategorised",
            "category": _clean(r[CAT]) or "",
            "mrp": mrp,
            "image_filename": _clean(r[SITE]),
            "image_path": None,
        })
    return out

def _index_images(images_root):
    idx = {}
    for dirpath, _, files in os.walk(images_root):
        for f in files:
            if f.lower().endswith((".png", ".jpg", ".jpeg", ".webp")):
                idx.setdefault(f, os.path.join(dirpath, f))
    return idx

def resolve_images(products, images_root):
    idx = _index_images(images_root)
    included, no_image, broken = [], [], []
    used = set()
    for p in products:
        fn = p["image_filename"]
        if not fn:
            no_image.append(p); continue
        path = idx.get(fn)
        if not path:
            broken.append(p); continue
        p["image_path"] = path
        used.add(fn)
        included.append(p)
    orphans = sorted(set(idx) - used)
    return included, no_image, broken, orphans

SIGNATURE_PREFIXES = ["Zenith", "Opell Prima", "Para"]

def order_collections(products):
    # preserve first-appearance order of collections
    seen = []
    buckets = {}
    for p in products:
        c = p["collection"]
        if c not in buckets:
            buckets[c] = []; seen.append(c)
        buckets[c].append(p)

    def rank(name):
        for i, pre in enumerate(SIGNATURE_PREFIXES):
            if name.startswith(pre):
                return (i, seen.index(name))
        return (len(SIGNATURE_PREFIXES), seen.index(name))

    ordered_names = sorted(seen, key=rank)
    return [(name, buckets[name]) for name in ordered_names]

def paginate(groups):
    per_page = COLS * ROWS
    pages = []
    for name, items in groups:
        for i in range(0, len(items), per_page):
            pages.append((name, items[i:i+per_page]))
    return pages

def write_report(report_path, total, included, no_image, broken, price_on_request, orphans):
    Path(report_path).parent.mkdir(parents=True, exist_ok=True)
    L = []
    L.append("# Catalogue Build Report\n")
    L.append(f"- Total rows: {total}")
    L.append(f"- Included: {included}")
    L.append(f"- Omitted (no image): {len(no_image)}")
    L.append(f"- Omitted (broken image ref): {len(broken)}")
    L.append(f"- Price on request: {len(price_on_request)}")
    L.append(f"- Orphan photos (on disk, no included row): {len(orphans)}\n")

    def section(title, rows, fmt):
        L.append(f"\n## {title} ({len(rows)})\n")
        if not rows: L.append("_none_"); return
        for r in rows: L.append(fmt(r))

    section("Omitted — no image", no_image, lambda r: f"- {r['item_code']} — {r['name']} ({r['collection']})")
    section("Omitted — broken image reference", broken, lambda r: f"- {r['item_code']} — {r['name']} → {r.get('image_filename')}")
    section("Price on request", price_on_request, lambda r: f"- {r['item_code']} — {r['name']} ({r['collection']})")
    section("Orphan photos", [{"f":o} for o in orphans], lambda r: f"- {r['f']}")

    Path(report_path).write_text("\n".join(L), encoding="utf-8")
    print(f"Report: {report_path}  |  included={included} no_image={len(no_image)} broken={len(broken)} price_req={len(price_on_request)} orphans={len(orphans)}")

from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.utils import ImageReader
from PIL import Image
import io

def _draw_logo(c, x, y, scale=1.0):
    """Minimal IPM wordmark: 'IPM' + 4 right bars + 'BATH FITTINGS'."""
    c.setFillColor(HexColor("#111111"))
    c.setFont("SegoeUI-Bold", 26*scale)
    c.drawString(x, y, "IPM")
    bx = x + 52*scale
    c.setLineWidth(2.2*scale); c.setStrokeColor(HexColor("#111111"))
    for i, w in enumerate([26, 22, 18, 14]):
        yy = y + 18*scale - i*5.5*scale
        c.line(bx + (26-w)*scale, yy, bx + 26*scale, yy)
    c.setFont("SegoeUI", 6.5*scale)
    c.drawString(x, y - 9*scale, "B A T H   F I T T I N G S")

def _desaturate(path):
    im = Image.open(path).convert("RGB")
    g = im.convert("L").convert("RGB")
    buf = io.BytesIO(); g.save(buf, format="JPEG", quality=88); buf.seek(0)
    return ImageReader(buf)

# Cache downscaled product images by absolute source path so the same file is
# only decoded/resized once even if it appears on multiple pages.
_PRODUCT_IMG_CACHE = {}
PRODUCT_IMG_MAX_PX = 320

def _load_product_image(path):
    """Return (ImageReader, (w, h)) for a downscaled, transparency-preserving copy
    of the product photo. Cached by absolute path."""
    key = os.path.abspath(path)
    cached = _PRODUCT_IMG_CACHE.get(key)
    if cached is not None:
        return cached
    im = Image.open(path)
    im = im.convert("RGBA")  # keep transparency
    im.thumbnail((PRODUCT_IMG_MAX_PX, PRODUCT_IMG_MAX_PX), Image.LANCZOS)
    buf = io.BytesIO(); im.save(buf, format="PNG", optimize=True); buf.seek(0)
    reader = ImageReader(buf)
    result = (reader, im.size)
    _PRODUCT_IMG_CACHE[key] = result
    return result

def draw_cover(c, hero_path, wef="01.04.2026"):
    if hero_path and os.path.exists(hero_path):
        c.drawImage(_desaturate(hero_path), 0, 0, PAGE_W, PAGE_H, preserveAspectRatio=False, mask=None)
        c.setFillColor(HexColor("#000000")); c.setFillAlpha(0.28)
        c.rect(0, 0, PAGE_W, PAGE_H, fill=1, stroke=0); c.setFillAlpha(1)
    else:
        c.setFillColor(HexColor("#2b2b2b")); c.rect(0,0,PAGE_W,PAGE_H,fill=1,stroke=0)
    c.saveState(); c.translate(PAGE_W-150, PAGE_H-80)
    c.setFillColor(HexColor("#ffffff")); c.setStrokeColor(HexColor("#ffffff"))
    c.setFont("SegoeUI-Bold", 26); c.drawString(0,0,"IPM")
    for i,w in enumerate([26,22,18,14]):
        yy = 18 - i*5.5; c.setLineWidth(2.2); c.line((26-w)+52, yy+0, 52+26, yy)
    c.setFont("SegoeUI", 6.5); c.drawString(0,-9,"B A T H   F I T T I N G S"); c.restoreState()
    c.setFillColor(HexColor("#ffffff"))
    c.setFont("SegoeUI-Semibold", 20); c.drawCentredString(PAGE_W/2, 118, "F U L F I L L I N G   Y O U R   B A T H I N G   D E S I R E S")
    c.setFont("SegoeUI-Bold", 26); c.drawCentredString(PAGE_W/2, 70, "INDIA MRP LIST")
    c.setFont("SegoeUI-Semibold", 11); c.drawCentredString(PAGE_W/2, 48, f"W.E.F : {wef}")
    c.showPage()

def draw_header(c, collection):
    c.setFillColor(TEAL); c.rect(0, PAGE_H-HEADER_H, PAGE_W, HEADER_H, fill=1, stroke=0)
    # white rounded plate cradling logo (left)
    c.setFillColor(HexColor("#ffffff"))
    c.roundRect(-20, PAGE_H-HEADER_H+8, 300, HEADER_H-16, 18, fill=1, stroke=0)
    _draw_logo(c, 34, PAGE_H-HEADER_H+40, scale=1.0)
    # gray tab (right) with collection name
    tab_x = PAGE_W*0.42
    c.setFillColor(GRAY_TAB); c.roundRect(tab_x, PAGE_H-HEADER_H+18, PAGE_W-tab_x-MARGIN_X, HEADER_H-36, 10, fill=1, stroke=0)
    c.setFillColor(GRAY_TAB_TEXT); c.setFont("SegoeUI-Bold", 18)
    c.drawRightString(PAGE_W-MARGIN_X-14, PAGE_H-HEADER_H+40, collection.upper())

def draw_footer(c, page_no):
    c.setFillColor(TEAL); c.rect(0, 0, PAGE_W, FOOTER_H, fill=1, stroke=0)
    c.setFillColor(HexColor("#ffffff")); c.setFont("SegoeUI-Bold", 11)
    c.drawRightString(PAGE_W-MARGIN_X, 10, str(page_no))

def _fmt_price(mrp):
    return f"₹ : {mrp}/-" if mrp is not None else "Price on request"

def draw_cell(c, x, y, w, h, product):
    # code box (teal) top-left — auto-fit width to the code, min 42pt like the sample
    code = str(product["item_code"])
    code_font, code_size, code_pad = "SegoeUI-Bold", 9, 8
    box_w = max(42, c.stringWidth(code, code_font, code_size) + 2*code_pad)
    c.setFillColor(TEAL); c.rect(x, y+h-20, box_w, 16, fill=1, stroke=0)
    c.setFillColor(HexColor("#ffffff")); c.setFont(code_font, code_size)
    c.drawCentredString(x+box_w/2, y+h-16, code)
    # image area
    img_h = h*0.55
    if product.get("image_path") and os.path.exists(product["image_path"]):
        try:
            ir, (iw, ih) = _load_product_image(product["image_path"])
            box_w, box_h = w*0.8, img_h
            scale = min(box_w/iw, box_h/ih)
            dw, dh = iw*scale, ih*scale
            c.drawImage(ir, x+(w-dw)/2, y+h-24-dh, dw, dh, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    # name (uppercase, may wrap 2 lines)
    c.setFillColor(HexColor("#222222")); c.setFont("SegoeUI-Semibold", 8.5)
    name = product["name"].upper()
    ty = y + h - 24 - img_h - 8
    for line in _wrap(c, name, "SegoeUI-Semibold", 8.5, w-4)[:2]:
        c.drawString(x+2, ty, line); ty -= 11
    # price
    c.setFont("SegoeUI-Bold", 8.5); c.setFillColor(HexColor("#111111"))
    c.drawString(x+2, ty-1, _fmt_price(product.get("mrp")))

def _wrap(c, text, font, size, max_w):
    words = text.split(); lines=[]; cur=""
    for wd in words:
        t = (cur+" "+wd).strip()
        if c.stringWidth(t, font, size) <= max_w: cur=t
        else: lines.append(cur); cur=wd
    if cur: lines.append(cur)
    return lines or [""]

def build_pdf(groups, out_path, hero_path):
    register_fonts()
    c = rl_canvas.Canvas(str(out_path), pagesize=(PAGE_W, PAGE_H))
    draw_cover(c, hero_path)
    pages = paginate(groups)
    grid_top = PAGE_H - HEADER_H - 8
    grid_bottom = FOOTER_H + 8
    cell_w = (PAGE_W - 2*MARGIN_X) / COLS
    cell_h = (grid_top - grid_bottom) / ROWS
    for pno, (coll, items) in enumerate(pages, start=1):
        draw_header(c, coll)
        for idx, product in enumerate(items):
            r, col = divmod(idx, COLS)
            x = MARGIN_X + col*cell_w
            y = grid_top - (r+1)*cell_h
            draw_cell(c, x+6, y+4, cell_w-12, cell_h-8, product)
        draw_footer(c, pno)
        c.showPage()
    c.save()

import argparse

def main(argv=None):
    ap = argparse.ArgumentParser(description="Build IPM catalogue PDF")
    ap.add_argument("--xlsx", default="ITEM MASTER FOR WEBSITE.xlsx")
    ap.add_argument("--images", default="images/products")
    ap.add_argument("--hero", default="images/home/hero.jpg")
    ap.add_argument("--out", default="IPM Catalogue.pdf")
    ap.add_argument("--report", default="reports/catalogue-build.md")
    args = ap.parse_args(argv)

    products = load_products(args.xlsx)
    included, no_image, broken, orphans = resolve_images(products, args.images)
    price_req = [p for p in included if p["mrp"] is None]
    groups = order_collections(included)
    build_pdf(groups, args.out, args.hero)
    write_report(args.report, total=len(products), included=len(included),
                 no_image=no_image, broken=broken, price_on_request=price_req, orphans=orphans)
    print(f"PDF: {args.out}  ({len(included)} products, {len(groups)} collections)")

if __name__ == "__main__":
    main()
